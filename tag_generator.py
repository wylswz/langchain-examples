"""标签生成器 - 使用 LLM 为用户评价生成标签，区分满意/不满意，输出 0/1 关联矩阵"""

import csv
import os
import re
import time
import warnings
from pathlib import Path

from openpyxl import load_workbook
from pydantic import BaseModel, Field
from langchain_openai import ChatOpenAI

warnings.filterwarnings("ignore", message=".*PydanticSerializationUnexpectedValue.*")

# ---------------------------------------------------------------------------
# 标签分类指引 —— 用于 LLM prompt
# ---------------------------------------------------------------------------

TAG_GUIDE = """
以下是标签分类指引，请严格按照这些标签名进行归类：

- 华为生态：手表、耳机、车机、平板互联、多设备协同等
- 性能：芯片、处理器、整机性能、发热等
- 操作系统：系统流畅性、操作交互、UI 设计、系统稳定性、系统更新等
- 三方应用：App 适配、软件兼容性、应用数量、功能缺失等
- 影像能力：拍照、录像、摄影、夜景、防抖，红枫这类也属于影像能力
- 外观手感：外观、颜色、手感、重量、尺寸、做工等
- 外观设计：注意和上条没有重复，是上条标签的展开，可以当作两个标签
- 手感：包括尺寸、重量、握持手感等
- 续航充电：电池续航、充电速度、电量等
- 信号通信：信号强度、网络稳定性、通话质量、WiFi等
- 屏幕显示：屏幕效果、刷新率、亮度、色彩、护眼等
- 品牌价值：品牌信任、国产支持、品质、售后服务等
- AI功能：小艺助手、AI 能力、智能功能等
- 音质音效：扬声器、音质、立体声、外放等
- 安全隐私：安全性、隐私保护等
- 价格性价比：价格、性价比、定价等
- 整体满意：只是综合泛泛的评价，没有多余信息
- 屏幕
- 服务

如果陈述中涉及的方面不在以上列表中，可以新增标签，但名称风格需保持一致（2-5个字的简短名词短语）。
同一句话中的同类内容可以拆分成多个标签（如"尺寸大，手机重"归为"外观手感"和"手感"）。
""".strip()


# ---------------------------------------------------------------------------
# Pydantic 结构化输出
# ---------------------------------------------------------------------------


class StatementAnalysis(BaseModel):
    """对一条用户评价的分析结果"""

    satisfaction_tags: list[str] = Field(
        default_factory=list,
        description="该评价中满意/正面方面对应的标签列表",
    )
    dissatisfaction_tags: list[str] = Field(
        default_factory=list,
        description="该评价中不满意/负面方面对应的标签列表",
    )


# ---------------------------------------------------------------------------
# File I/O
# ---------------------------------------------------------------------------


def load_statements(file_path: str, statement_column: int = 1) -> tuple[list[str], str, object]:
    """从 CSV / XLSX 中读取陈述列表。"""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        wb = load_workbook(file_path)
        ws = wb.active
        statements = []
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=statement_column).value
            if val:
                statements.append(str(val).strip())
        return statements, "excel", wb

    if suffix == ".csv":
        with open(file_path, encoding="utf-8") as f:
            rows = list(csv.reader(f))
        if not rows:
            raise ValueError("CSV 文件为空")
        idx = statement_column - 1
        statements = [
            str(row[idx]).strip()
            for row in rows[1:]
            if row and len(row) > idx and row[idx].strip()
        ]
        return statements, "csv", rows

    raise ValueError(f"不支持的文件格式: {suffix}")


def save_matrix_csv(
    save_path: str,
    statements: list[str],
    all_tags: list[str],
    tag_map: list[set[str]],
):
    """将 0/1 矩阵写入 CSV。"""
    header = ["陈述"] + all_tags
    with open(save_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        for stmt, tags_for_row in zip(statements, tag_map):
            row = [stmt] + ["1" if tag in tags_for_row else "0" for tag in all_tags]
            w.writerow(row)


def save_matrix_excel(
    save_path: str,
    statements: list[str],
    all_tags: list[str],
    tag_map: list[set[str]],
):
    """将 0/1 矩阵写入 XLSX。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["陈述"] + all_tags)
    for stmt, tags_for_row in zip(statements, tag_map):
        row = [stmt] + [1 if tag in tags_for_row else 0 for tag in all_tags]
        ws.append(row)
    wb.save(save_path)


def load_batch_csv(
    sat_path: str, dissat_path: str
) -> tuple[list[set[str]], list[set[str]], list[str], list[str]]:
    """从已保存的批次 CSV 加载 tag_map。返回 (sat_map, dissat_map, sat_tags, dissat_tags)。"""
    sat_map: list[set[str]] = []
    dissat_map: list[set[str]] = []

    with open(sat_path, encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        return sat_map, dissat_map, [], []
    sat_tags = rows[0][1:]
    for row in rows[1:]:
        tags_row = {sat_tags[j] for j in range(len(sat_tags)) if j + 1 < len(row) and row[j + 1].strip() == "1"}
        sat_map.append(tags_row)

    with open(dissat_path, encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        return sat_map, dissat_map, sat_tags, []
    dissat_tags = rows[0][1:]
    for row in rows[1:]:
        tags_row = {dissat_tags[j] for j in range(len(dissat_tags)) if j + 1 < len(row) and row[j + 1].strip() == "1"}
        dissat_map.append(tags_row)

    return sat_map, dissat_map, sat_tags, dissat_tags


# ---------------------------------------------------------------------------
# LLM
# ---------------------------------------------------------------------------


def create_llm(
    model: str | None = None,
    base_url: str | None = None,
    api_key: str | None = None,
    temperature: float = 0,
):
    """创建配置了结构化输出的 LLM。"""
    model = model or os.getenv("OPENAI_MODEL", "gpt-4o-mini")
    base_url = base_url or os.getenv("OPENAI_BASE_URL")
    api_key = api_key or os.getenv("OPENAI_API_KEY")

    kwargs: dict = {"model": model, "temperature": temperature}
    if base_url:
        kwargs["base_url"] = base_url
    if api_key:
        kwargs["api_key"] = api_key

    return ChatOpenAI(**kwargs).with_structured_output(StatementAnalysis)


SYSTEM_PROMPT = f"""你是一个专业的用户评价分析助手。你的任务是分析用户对手机产品的评价，提取其中的**满意点**和**不满意点**，并为每个方面分配标签。

{TAG_GUIDE}

规则：
1. 一条评价可能同时包含满意和不满意的内容，请分别归类。
2. 只选择真正相关的标签，不要过度标记。
3. 同一方面只标记一次，不要重复。
4. 如果评价中没有明确的满意或不满意内容，对应列表留空。

标签（包括但不限于）

如果没有统计到的地方，可以自行补充标签
"""


def sanitize_statement(s: str, max_len: int = 6000) -> str:
    """清理陈述文本，避免 JSON/API 请求异常。"""
    if not s:
        return ""
    # 移除控制字符（保留 \n \t \r）
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", s)
    # 替换异常 Unicode（如替换字符 U+FFFD）
    s = s.encode("utf-8", errors="replace").decode("utf-8")
    return s[:max_len] if len(s) > max_len else s


def analyze_statement(llm, statement: str, max_retries: int = 3) -> StatementAnalysis:
    """分析一条用户评价，返回满意/不满意标签。"""
    clean = sanitize_statement(statement)
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": f"请分析以下用户评价：\n\n{clean}"},
    ]
    for attempt in range(max_retries):
        try:
            return llm.invoke(messages)
        except Exception as e:
            if attempt < max_retries - 1:
                wait = 2 ** attempt
                print(f"  ! API 错误，{wait}s 后重试: {e}", flush=True)
                time.sleep(wait)
            else:
                raise


# ---------------------------------------------------------------------------
# Core pipeline
# ---------------------------------------------------------------------------


def process_file(
    file_path: str,
    output_dir: str = "",
    statement_column: int = 1,
    output_format: str = "",
    batch_size: int = 100,
    start_from: int = 1,
):
    """处理文件：分析评价 -> 拆分满意/不满意 -> 输出 0/1 矩阵文件。

    每处理 batch_size 条输出一批中间文件，全部完成后输出汇总文件。

    Args:
        file_path: 输入文件路径
        output_dir: 输出目录（默认与输入文件同目录）
        statement_column: 陈述所在列号（从 1 开始）
        output_format: 输出格式 "csv" 或 "xlsx"（默认与输入相同）
        batch_size: 每批条数，每批输出一次中间文件，默认 100
        start_from: 从第几条开始处理（用于断点续跑），默认 1
    """
    print(f"加载文件: {file_path}", flush=True)
    statements, file_type, _ = load_statements(file_path, statement_column)
    print(f"共 {len(statements)} 条陈述，每 {batch_size} 条输出一批", flush=True)
    if start_from > 1:
        print(f"从第 {start_from} 条开始处理（跳过前 {start_from - 1} 条）", flush=True)
    print("-" * 50, flush=True)

    src = Path(file_path)
    out_dir = Path(output_dir) if output_dir else src.parent
    fmt = output_format or ("xlsx" if file_type == "excel" else "csv")
    save_fn = save_matrix_excel if fmt == "xlsx" else save_matrix_csv

    sat_tags_per_row: list[set[str]] = []
    dissat_tags_per_row: list[set[str]] = []
    all_sat_tags: set[str] = set()
    all_dissat_tags: set[str] = set()

    # 从已有批次加载前 start_from-1 条结果
    if start_from > 1 and fmt == "csv":
        for batch_start in range(1, start_from, batch_size):
            batch_end = min(batch_start + batch_size - 1, start_from - 1)
            sp = out_dir / f"{src.stem}_满意_{batch_start:04d}-{batch_end:04d}.{fmt}"
            dp = out_dir / f"{src.stem}_不满意_{batch_start:04d}-{batch_end:04d}.{fmt}"
            if sp.exists() and dp.exists():
                ps, pd, pst, pdt = load_batch_csv(str(sp), str(dp))
                sat_tags_per_row.extend(ps)
                dissat_tags_per_row.extend(pd)
                all_sat_tags.update(pst)
                all_dissat_tags.update(pdt)
        if len(sat_tags_per_row) < start_from - 1:
            raise FileNotFoundError(
                f"未找到前 {start_from - 1} 条的批次文件（需 {src.stem}_满意_0001-0100.csv 等），请先完成前 {start_from - 1} 条处理"
            )
        print(f"已加载前 {len(sat_tags_per_row)} 条结果", flush=True)

    llm = create_llm()
    batch_sat_tags: set[str] = set()
    batch_dissat_tags: set[str] = set()

    for i in range(start_from - 1, len(statements)):
        stmt = statements[i]
        preview = stmt.replace("\n", " ")[:60]
        print(f"[{i + 1}/{len(statements)}] {preview}...", flush=True)

        result = analyze_statement(llm, stmt)

        sat = set(t for t in result.satisfaction_tags if t and str(t).strip())
        dissat = set(t for t in result.dissatisfaction_tags if t and str(t).strip())

        sat_tags_per_row.append(sat)
        dissat_tags_per_row.append(dissat)
        all_sat_tags.update(sat)
        all_dissat_tags.update(dissat)
        batch_sat_tags.update(sat)
        batch_dissat_tags.update(dissat)

        if sat:
            print(f"  ✓ 满意: {sorted(sat)}", flush=True)
        if dissat:
            print(f"  ✗ 不满意: {sorted(dissat)}", flush=True)

        # 每 batch_size 条输出一批（按 1-based 序号）
        n = i + 1
        if n >= start_from and (n - start_from + 1) % batch_size == 0:
            start = n - batch_size + 1
            end = n
            # 确保 start 不小于 1（续跑时可能不足整批）
            start = max(start, start_from)
            batch_stmts = statements[start - 1 : end]
            batch_sat_map = sat_tags_per_row[start - 1 : end]
            batch_dissat_map = dissat_tags_per_row[start - 1 : end]
            sorted_batch_sat = sorted(batch_sat_tags)
            sorted_batch_dissat = sorted(batch_dissat_tags)

            sat_batch_path = out_dir / f"{src.stem}_满意_{start:04d}-{end:04d}.{fmt}"
            dissat_batch_path = out_dir / f"{src.stem}_不满意_{start:04d}-{end:04d}.{fmt}"

            save_fn(str(sat_batch_path), batch_stmts, sorted_batch_sat, batch_sat_map)
            save_fn(str(dissat_batch_path), batch_stmts, sorted_batch_dissat, batch_dissat_map)

            print(f"  >> 已输出第 {n // batch_size} 批: {sat_batch_path.name}", flush=True)

            batch_sat_tags = set()
            batch_dissat_tags = set()

    # 最后不足 batch_size 的余数也输出一批
    last_batch_start = (len(statements) - 1) // batch_size * batch_size + 1
    if last_batch_start >= start_from and last_batch_start <= len(statements):
        start = max(last_batch_start, start_from)
        end = len(statements)
        batch_stmts = statements[start - 1 : end]
        batch_sat_map = sat_tags_per_row[start - 1 : end]
        batch_dissat_map = dissat_tags_per_row[start - 1 : end]
        if batch_sat_tags or batch_dissat_tags:
            sorted_batch_sat = sorted(batch_sat_tags)
            sorted_batch_dissat = sorted(batch_dissat_tags)
            sat_batch_path = out_dir / f"{src.stem}_满意_{start:04d}-{end:04d}.{fmt}"
            dissat_batch_path = out_dir / f"{src.stem}_不满意_{start:04d}-{end:04d}.{fmt}"
            save_fn(str(sat_batch_path), batch_stmts, sorted_batch_sat, batch_sat_map)
            save_fn(str(dissat_batch_path), batch_stmts, sorted_batch_dissat, batch_dissat_map)
            print(f"  >> 已输出最后一批: {sat_batch_path.name}", flush=True)

    sorted_sat_tags = sorted(all_sat_tags)
    sorted_dissat_tags = sorted(all_dissat_tags)

    print("-" * 50, flush=True)
    print(f"满意标签 ({len(sorted_sat_tags)}): {sorted_sat_tags}", flush=True)
    print(f"不满意标签 ({len(sorted_dissat_tags)}): {sorted_dissat_tags}", flush=True)

    # 输出汇总文件
    sat_path = str(out_dir / f"{src.stem}_满意.{fmt}")
    dissat_path = str(out_dir / f"{src.stem}_不满意.{fmt}")

    save_fn(sat_path, statements, sorted_sat_tags, sat_tags_per_row)
    print(f"满意汇总已保存: {sat_path}", flush=True)

    save_fn(dissat_path, statements, sorted_dissat_tags, dissat_tags_per_row)
    print(f"不满意汇总已保存: {dissat_path}", flush=True)

    return sat_path, dissat_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="标签生成器 - 分析用户评价，生成满意/不满意标签矩阵"
    )
    parser.add_argument("file_path", help="输入文件路径（.csv / .xlsx / .xls）")
    parser.add_argument("-o", "--output-dir", default="", help="输出目录（默认与输入文件同目录）")
    parser.add_argument(
        "-s", "--statement-column", type=int, default=1,
        help="陈述所在列号（从 1 开始），默认 1",
    )
    parser.add_argument(
        "-f", "--format", default="", choices=["csv", "xlsx", ""],
        help="输出格式（默认与输入相同）",
    )
    parser.add_argument(
        "-b", "--batch-size", type=int, default=100,
        help="每批条数，每批输出一次中间文件，默认 100",
    )
    parser.add_argument(
        "--start", type=int, default=1,
        help="从第几条开始处理（断点续跑），默认 1",
    )
    args = parser.parse_args()
    process_file(
        args.file_path,
        args.output_dir,
        args.statement_column,
        args.format,
        args.batch_size,
        args.start,
    )


if __name__ == "__main__":
    main()
