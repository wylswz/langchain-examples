"""生成 tagger 测试数据"""

import csv
from openpyxl import Workbook


# 定义标签
TAGS = ["相机", "性能", "屏幕", "电池", "音质", "做工", "系统", "发热", "信号"]

# 定义测试陈述（带有 ID 和来源信息）
STATEMENTS = [
    ("R001", "用户A", "在人像拍摄的时候，边缘解析力明显下降"),
    ("R002", "用户B", "玩大型游戏时手机发烫严重，帧率不稳定"),
    ("R003", "用户C", "屏幕色彩很鲜艳，但是在户外阳光下看不清"),
    ("R004", "用户A", "电池续航非常给力，充电速度也很快"),
    ("R005", "用户D", "外放音质一般，低音不够有力"),
    ("R006", "用户B", "金属边框做工精细，手感很好"),
    ("R007", "用户E", "系统操作流畅，但广告太多了"),
    ("R008", "用户C", "夜景模式拍摄效果很惊艳"),
    ("R009", "用户F", "信号不太稳定，地铁里经常断网"),
    ("R010", "用户A", "充电的时候机器特别烫"),
    ("R011", "用户G", "扬声器音量很大，立体声效果不错"),
    ("R012", "用户B", "屏幕刷新率高，滑动很顺滑"),
    ("R013", "用户H", "后盖塑料感太强，按压有声响"),
    ("R014", "用户D", "拍视频的时候防抖效果很好"),
    ("R015", "用户E", "开机启动太慢，系统优化不到位"),
]


def create_excel_test_data(filename="test_data.xlsx"):
    """创建 Excel 测试数据（简单格式：陈述在第1列，标签从第2列开始）"""
    wb = Workbook()
    ws = wb.active

    # 写入表头
    ws.cell(row=1, column=1, value="陈述")
    for i, tag in enumerate(TAGS):
        ws.cell(row=1, column=i + 2, value=tag)

    # 写入陈述（只取陈述内容，标签列留空）
    for i, (_, _, statement) in enumerate(STATEMENTS):
        ws.cell(row=i + 2, column=1, value=statement)

    wb.save(filename)
    print(f"Excel 测试数据已创建: {filename}")
    print(f"  格式: 陈述列=1")


def create_csv_test_data(filename="test_data.csv"):
    """创建 CSV 测试数据（简单格式：陈述在第1列，标签从第2列开始）"""
    with open(filename, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)

        # 写入表头
        header = ["陈述"] + TAGS
        writer.writerow(header)

        # 写入陈述（只取陈述内容，标签列留空）
        for _, _, statement in STATEMENTS:
            row = [statement] + [""] * len(TAGS)
            writer.writerow(row)

    print(f"CSV 测试数据已创建: {filename}")
    print(f"  格式: 陈述列=1")


def create_excel_test_data_extended(filename="test_data_extended.xlsx"):
    """创建 Excel 测试数据（扩展格式：包含 ID、来源等只读列）

    列结构: ID(1) | 来源(2) | 陈述(3) | 标签...(4+)
    """
    wb = Workbook()
    ws = wb.active

    # 写入表头
    ws.cell(row=1, column=1, value="ID")
    ws.cell(row=1, column=2, value="来源")
    ws.cell(row=1, column=3, value="陈述")
    for i, tag in enumerate(TAGS):
        ws.cell(row=1, column=i + 4, value=tag)

    # 写入数据
    for i, (id_, source, statement) in enumerate(STATEMENTS):
        ws.cell(row=i + 2, column=1, value=id_)
        ws.cell(row=i + 2, column=2, value=source)
        ws.cell(row=i + 2, column=3, value=statement)

    wb.save(filename)
    print(f"Excel 扩展测试数据已创建: {filename}")
    print(f"  格式: ID列=1, 来源列=2, 陈述列=3（标签从第4列开始）")
    print(f"  使用时: python tagger.py {filename} -s 3")


def create_csv_test_data_extended(filename="test_data_extended.csv"):
    """创建 CSV 测试数据（扩展格式：包含 ID、来源等只读列）

    列结构: ID(1) | 来源(2) | 陈述(3) | 标签...(4+)
    """
    with open(filename, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)

        # 写入表头
        header = ["ID", "来源", "陈述"] + TAGS
        writer.writerow(header)

        # 写入数据
        for id_, source, statement in STATEMENTS:
            row = [id_, source, statement] + [""] * len(TAGS)
            writer.writerow(row)

    print(f"CSV 扩展测试数据已创建: {filename}")
    print(f"  格式: ID列=1, 来源列=2, 陈述列=3（标签从第4列开始）")
    print(f"  使用时: python tagger.py {filename} -s 3")


if __name__ == "__main__":
    # 创建简单格式的测试数据
    create_excel_test_data()
    create_csv_test_data()

    print()

    # 创建扩展格式的测试数据（带有只读列）
    create_excel_test_data_extended()
    create_csv_test_data_extended()

    print("\n" + "=" * 50)
    print("测试数据创建完成！")
    print(f"共 {len(TAGS)} 个标签: {TAGS}")
    print(f"共 {len(STATEMENTS)} 条陈述")
    print("\n使用示例:")
    print("  # 简单格式（默认参数，陈述在第1列）")
    print("  python tagger.py test_data.csv")
    print("  python tagger.py test_data.xlsx -o output.xlsx")
    print()
    print("  # 扩展格式（陈述在第3列，前两列只读）")
    print("  python tagger.py test_data_extended.csv -s 3")
    print("  python tagger.py test_data_extended.xlsx -s 3 -o output.xlsx")
