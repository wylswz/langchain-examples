"""智能标签助手 - 使用 LLM 结构化输出为陈述打标签"""

import csv
from pathlib import Path

from openpyxl import load_workbook, Workbook
from pydantic import BaseModel, Field
from langchain_openai import ChatOpenAI


class TagResult(BaseModel):
    """标签识别结果"""

    tags: list[str] = Field(
        description="识别出的相关标签列表，只能从给定的标签列表中选择"
    )


def load_file(file_path: str, statement_column: int = 1):
    """加载文件，返回表头、陈述列表和原始数据

    Args:
        file_path: 文件路径（支持 .xlsx, .xls, .csv）
        statement_column: 陈述所在的列号（从 1 开始），标签从下一列开始

    Returns:
        (tags, statements, file_type, raw_data)
        - tags: 标签列表
        - statements: 陈述列表
        - file_type: 'excel' 或 'csv'
        - raw_data: 原始数据（workbook 或 csv rows）
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    suffix = path.suffix.lower()
    tag_start_column = statement_column + 1

    if suffix in [".xlsx", ".xls"]:
        wb = load_workbook(file_path)
        ws = wb.active

        # 读取标签（从 tag_start_column 开始）
        tags = []
        for col in range(tag_start_column, ws.max_column + 1):
            tag = ws.cell(row=1, column=col).value
            if tag:
                tags.append(str(tag))

        # 读取陈述
        statements = []
        for row in range(2, ws.max_row + 1):
            statement = ws.cell(row=row, column=statement_column).value
            if statement:
                statements.append(str(statement))

        return tags, statements, "excel", wb

    elif suffix == ".csv":
        with open(file_path, "r", encoding="utf-8") as f:
            rows = list(csv.reader(f))

        if not rows:
            raise ValueError("CSV 文件为空")

        # 读取标签
        header = rows[0]
        tag_start_idx = tag_start_column - 1
        tags = [str(tag) for tag in header[tag_start_idx:] if tag]

        # 读取陈述
        statement_idx = statement_column - 1
        statements = []
        for row in rows[1:]:
            if row and len(row) > statement_idx and row[statement_idx]:
                statements.append(str(row[statement_idx]))

        return tags, statements, "csv", rows

    else:
        raise ValueError(f"不支持的文件格式: {suffix}")


def mark_tags_excel(
    wb,
    row_index: int,
    tags_to_mark: list[str],
    all_tags: list[str],
    tag_start_column: int,
):
    """在 Excel 中标记标签"""
    ws = wb.active
    excel_row = row_index + 2  # +1 表头 +1 因为从 1 开始

    for tag in tags_to_mark:
        if tag in all_tags:
            col_index = all_tags.index(tag) + tag_start_column
            ws.cell(row=excel_row, column=col_index, value=1)


def mark_tags_csv(
    rows: list,
    row_index: int,
    tags_to_mark: list[str],
    all_tags: list[str],
    tag_start_column: int,
):
    """在 CSV 数据中标记标签"""
    csv_row = row_index + 1  # +1 因为表头占了第 0 行
    tag_start_idx = tag_start_column - 1

    # 确保行有足够的列
    while len(rows[csv_row]) < tag_start_idx + len(all_tags):
        rows[csv_row].append("")

    for tag in tags_to_mark:
        if tag in all_tags:
            col_index = all_tags.index(tag) + tag_start_idx
            rows[csv_row][col_index] = "1"


def save_file(file_path: str, file_type: str, raw_data, output_path: str = ""):
    """保存文件"""
    save_path = output_path if output_path else file_path

    if file_type == "excel":
        raw_data.save(save_path)
    elif file_type == "csv":
        with open(save_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(raw_data)

    return save_path


def create_tagger_llm():
    """创建用于标签识别的 LLM"""
    llm = ChatOpenAI(model="gpt-4o-mini", temperature=0)
    return llm.with_structured_output(TagResult)


def tag_statement(llm, statement: str, available_tags: list[str]) -> list[str]:
    """使用 LLM 为单条陈述打标签

    Args:
        llm: 配置了结构化输出的 LLM
        statement: 要分析的陈述
        available_tags: 可用的标签列表

    Returns:
        识别出的标签列表
    """
    prompt = f"""分析以下陈述，判断它与哪些标签相关。

可用标签：{available_tags}

陈述：{statement}

请从可用标签中选择所有相关的标签。只选择真正相关的标签，不要过度标记。"""

    result: TagResult = llm.invoke(prompt)

    # 过滤掉不在可用标签中的标签
    valid_tags = [tag for tag in result.tags if tag in available_tags]
    return valid_tags


def process_file(file_path: str, output_path: str = "", statement_column: int = 1):
    """处理文件，为所有陈述打标签

    Args:
        file_path: 输入文件路径
        output_path: 输出文件路径（可选，默认覆盖原文件）
        statement_column: 陈述所在的列号（从 1 开始）
    """
    print(f"加载文件: {file_path}")
    tags, statements, file_type, raw_data = load_file(file_path, statement_column)
    tag_start_column = statement_column + 1

    print(f"文件类型: {file_type}")
    print(f"陈述列: {statement_column}，标签从第 {tag_start_column} 列开始")
    print(f"共 {len(tags)} 个标签: {tags}")
    print(f"共 {len(statements)} 条陈述待处理")
    print("-" * 50)

    # 创建 LLM
    llm = create_tagger_llm()

    # 逐行处理
    for i, statement in enumerate(statements):
        print(f"[{i + 1}/{len(statements)}] {statement[:50]}...")

        # 使用 LLM 识别标签
        matched_tags = tag_statement(llm, statement, tags)
        print(f"  -> 标签: {matched_tags}")

        # 标记标签
        if file_type == "excel":
            mark_tags_excel(raw_data, i, matched_tags, tags, tag_start_column)
        else:
            mark_tags_csv(raw_data, i, matched_tags, tags, tag_start_column)

    print("-" * 50)

    # 保存文件
    saved_path = save_file(file_path, file_type, raw_data, output_path)
    print(f"文件已保存: {saved_path}")

    return saved_path


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="智能标签助手 - 自动为陈述标记相关标签"
    )
    parser.add_argument("file_path", help="输入文件路径（.xlsx, .xls, .csv）")
    parser.add_argument(
        "-o", "--output", default="", help="输出文件路径（默认覆盖原文件）"
    )
    parser.add_argument(
        "-s",
        "--statement-column",
        type=int,
        default=1,
        help="陈述所在的列号（从 1 开始），默认为 1。标签从此列的下一列开始。",
    )

    args = parser.parse_args()

    process_file(args.file_path, args.output, args.statement_column)


if __name__ == "__main__":
    main()
